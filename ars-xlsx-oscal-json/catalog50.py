import json
import re
from datetime import datetime, timezone
from pathlib import Path
from uuid import uuid4

import click
import openpyxl
from trestle.oscal.catalog import Catalog, Control, Group, Model
from trestle.oscal.common import Link, Metadata, Part, Property

from complianceio.oscal.oscal import control_to_statement_id

import logging
import sys

# set to INFO or DEBUG for more verbose logging (ERROR for none)
logging.basicConfig(stream=sys.stdout, level=logging.ERROR)

'''
Convert ARS 5.0 spreadsheet into OSCAL code
SOURCE: https://security.cms.gov/library/cms-acceptable-risk-safeguards-ars
        ARS Full Element Principle_Single_Assessment_Current V5.01.xlsx

Todo:
* fix isoformat() -- missing sep='T'
* add debug arguments
* handle Redacted links
* add Organization Defined Parameters (ODPs)
'''


def create_groups(p, minimize):
    """
    Parse the spreadsheet to get groups by family.
    """
    with open(Path.cwd().joinpath('header50_map.json'), 'r') as f:
        c_map = json.load(f)
    wb = openpyxl.load_workbook(filename=p, data_only=True)
    if "Full" not in wb.sheetnames:
        print(f"Sheetname 'Full' not found in: {wb.sheetnames}")
        quit()
    ws = wb["Full"]
    # header = [cell for cell in ws['A2:XFD2'][0]
    #           if cell.value is not None and cell.value.strip() != '']
    groups = []
    family_id = None
    group = None
    controls = []
    for row in ws.iter_rows(min_row=3, max_col=19, max_row=2600):
        if row[0].value is not None:
            family = row[c_map.get('family')].value
            control_id = row[c_map.get('control_id')].value
            group = control_id[:2].lower()
            name = row[c_map.get('name')].value.strip()
            control_text = row[c_map.get('control')].value
            responsibility = row[c_map.get('responsiblity')].value
            related = row[c_map.get('responsiblity')].value
            reference = row[c_map.get('reference')].value
            baseline = row[c_map.get('baseline')].value
            priority = row[c_map.get('priority')].value

            # ignore duplicated parts (elements)
            if baseline is None:
                continue
            # check for HVA in element (line 332)
            if control_id == 'AT-02(01)a':
                continue
            # check for ac--09(02)
            control_id = control_id.replace('--', '-')
            # check for AU14-(01)
            control_id = re.sub(r'([A-Z][A-Z])([0-9][0-9])\-', r"\1-\2", control_id)

            if family_id is not None and family != family_id:
                g = add_group(group_id, family_id, controls)
                groups.append(g)
                controls = []

            if control_text is not None:
                cid = control_to_statement_id(control_id)
                parts = []
                part = parse_parts(control_text, cid)
                parts.append(part)

                imp = row[c_map.get('implementation')].value
                if imp and imp.strip():
                    pid = cid.replace('_smt', '_imp')
                    part = add_additional(pid, 'implementation', imp.strip())
                    parts.append(part)

                if not minimize:
                    hva = row[c_map.get('hva_standards')].value
                    if hva and hva.strip():
                        pid = cid.replace('_smt', '_hva')
                        part = add_additional(pid, 'hva', hva.strip())
                        parts.append(part)

                    prv = row[c_map.get('privacy_standards')].value
                    if prv and prv.strip():
                        pid = cid.replace('_smt', '_prv')
                        part = add_additional(pid, 'privacy', prv.strip())
                        parts.append(part)

                gdn = row[c_map.get('discussion')].value
                if gdn and gdn.strip():
                    pid = cid.replace('_smt', '_gdn')
                    part = add_additional(pid, 'guidance', gdn.strip())
                    parts.append(part)

                links = []
                related = row[c_map.get('related')].value
                if related and related.strip():
                    rlt = related.strip().replace(' ', ',').split(',')
                    for r in rlt:
                        if r[:4] != 'None':
                            link = r.strip()
                            if link:
                                # clean up Redacted links
                                if "Redacted" in link:
                                    continue
                                links.append(Link(
                                    href=f'#{link}',
                                    rel='related'
                                ))
                l = links if len(links) > 0 else None

                props = [
                    Property(
                        name='label',
                        value=control_id
                    ),
                    Property(
                        name='sort-id',
                        value=control_id.lower()
                    ),
                    Property(
                        name='responsibility',
                        value=str(responsibility.replace('\n', ',').strip().rstrip(','))
                    )
                ]
                if "Supplemental" in baseline:
                    control_class = "ARS-5.0-Supplemental"
                else:
                    control_class = "ARS-5.0-Mandatory"
                    props.append(
                        Property(
                            name='baseline',
                            value=str(baseline.replace('\n', ',').rstrip(','))
                        )
                    )
                if priority and priority.strip():
                    props.append(
                        Property(
                            name='priority',
                            value=priority.strip()
                        )
                    )
                controls.append(Control(
                    id=cid.split('_')[0],
                    class_=control_class,
                    title=name,
                    props=props,
                    links=l,
                    parts=parts,
                ))
        family_id = family
        group_id = group

    g = add_group(group_id, family, controls)
    groups.append(g)
    return groups


def add_group(group_id, family, controls):
    return Group(
        id=group_id,
        class_='family',
        title=family,
        controls=controls
    )


def parse_parts(text, cid):
    """
    This is very specific to the CMS spreadsheet.
    """
    pt = parse_control(text, cid)
    parts = []
    if pt:
        parts = create_parts(pt)

    part = Part(
        id=cid,
        name='statement',
        parts=parts,
    )

    return part


def parse_control(text, cid):
    """
    This is looking for a very specifically formatted control statement.
    Top level parts should start with (LETTER), for instance (a) or a.
    Secondary level should start with NUMBER. for instance 1. or i.
    Tertiary level should start with LETTER. for instance a.
    As an example:
    (a) This is the top level
    a. Could also be top-level
      a. Could also be top-level
      1. Could also be top-level
         1. Some secondary text
         (1) Also some secondary text
             a. And tertiary text.
    """
    lines = text.splitlines()
    parts = {cid.strip(): {'prose': None, 'children': {}}}
    pr = None
    fid = ''
    sid = ''
    for text in lines:
        firstdot = ''
        second = ''
        secondparen = ''
        third = ''
        first = re.search(r'^\s*\(([a-z])\)', text)             # |*(a)
        if not first:
            firstdot = re.search(r'^[a-z]\.', text)             # |a.
        if not first and not firstdot:
            second = re.search(r'^\s+[0-9i]+\.', text)          # |(a) 1.
            if second and not fid:
                firstdot = second                               # |  1.
        if not first and not firstdot and not second:
            secondparen = re.search(r'^\s+\(([0-9]+)\)', text)  # |(a) (1)
            if secondparen and not fid:
                first = secondparen                             # |  (1)
        if not first and not firstdot and not second and not secondparen:
            third = re.search(r'^\s+[a-z]\.', text)             # |(a) 1. a.
            if third and not sid:
                firstdot = third                                # |  a.
        text = text.strip()
        if first:
            f = first.group()
            logging.debug(f'1-"{f}"')
            fid = f.replace('(', '').replace(')', '')
            prose = text.strip(f'({f})').strip()
            parts[cid.strip()]['children'][fid.strip()] = {
                'prose': prose,
                'children': {},
            }
        elif firstdot:
            f = firstdot.group()
            logging.debug(f'1-"{f}"')
            fid = f.replace('.', '')
            prose = text.strip(f'{f}').strip()
            parts[cid.strip()]['children'][fid.strip()] = {
                'prose': prose,
                'children': {},
            }
        elif second:
            s = second.group()
            logging.debug(f'2-"{s}"')
            sid = s.replace('.', '')
            prose = text.strip(f'{s}').strip()
            parts[cid.strip()]['children'][fid.strip()]['children'][sid.strip()] = {
                'prose': prose,
                'children': {},
            }
        elif secondparen:
            s = secondparen.group()
            logging.debug(f'2-"{s}"')
            sid = s.replace('(', '').replace(')', '')
            prose = text.strip(f'{s}').strip()
            parts[cid.strip()]['children'][fid.strip()]['children'][sid.strip()] = {
                'prose': prose,
                'children': {},
            }
        elif third:
            t = third.group()
            logging.debug(f'3-"{t}"')
            tid = t.replace('.', '')
            prose = text.strip(f'{t}').strip()
            parts[cid.strip()]['children'][fid.strip()]\
                ['children'][sid.strip()]['children'][tid.strip()] = {'prose': prose}
        else:
            pr = f'{pr}\n{text}' if pr else text
            parts[cid.strip()]['prose'] = pr

    return parts


def create_parts(parts):
    part = []
    first = None
    secondary = None
    tertiary = None
    for k, pv in parts.items():
        if 'children' in pv:
            first = []
            for f, fv in pv.get('children').items():
                if 'children' in fv:
                    secondary = []
                    for s, sv in fv.get('children').items():
                        if 'children' in sv:
                            tertiary = []
                            for t, tv in sv.get('children').items():
                                logging.info(f'{k}.{f}.{s}.{t}')
                                tertiary.append(Part(
                                    id=f'{k}.{f}.{s}.{t}',
                                    name='item',
                                    props=[Property(
                                        name='label',
                                        value=t,
                                    )],
                                    prose=tv.get('prose')
                                ))
                        tp = tertiary if len(tertiary) > 0 else None
                        logging.info(f'{k}.{f}.{s}')
                        secondary.append(Part(
                            id=f'{k}.{f}.{s}',
                            name='item',
                            props=[Property(
                                name='label',
                                value=s,
                            )],
                            parts=tp,
                            prose=sv.get('prose'),
                        ))
                sp = secondary if len(secondary) > 0 else None
                logging.info(f'{k}.{f}')
                first.append(Part(
                    id=f'{k}.{f}',
                    name='item',
                    props=[Property(
                        name='label',
                        value=f,
                    )],
                    parts=sp,
                    prose=fv.get('prose'),
                ))
        fp = first if len(first) > 0 else None
        logging.info(f'{k}')
        part.append(Part(
            id=f'{k}',
            name='item',
            props=[Property(
                name='label',
                value=k,
            )],
            parts=fp,
            prose=pv.get('prose'),
        ))
    return part


def add_additional(pid, name, prose):
    return Part(
        id=pid,
        name=name,
        prose=prose,
    )


@click.command()
@click.option('-t', '--title', required=True)
@click.argument(
    'source',
    type=click.Path(exists=True, dir_okay=False, file_okay=True, resolve_path=True)

)
@click.option(
    "-a",
    "--ars_version",
    default="5.01",
    show_default=True,
    help="ARS version ",
)
@click.option(
    "-o",
    "--oscal_version",
    default="1.0.2",
    show_default=True,
    help="OSCAL version ",
)
@click.option(
    "-m",
    "--minimize",
    is_flag=True,
    help="Minimize output: no PII or HVA overlays.",
)
def main(title, source, ars_version, oscal_version, minimize):
    """
    Create an OSCAL Catalog from Excel (.xlsx) spreadsheet.
    Source: https://security.cms.gov/library/cms-acceptable-risk-safeguards-ars
    """
    p = Path(source)
    groups = create_groups(p, minimize)
    uuid = uuid4().urn


    # today = datetime.now(timezone(timedelta(hours=-6))).isoformat(sep='T')
    today = datetime.now().astimezone(timezone.utc).isoformat(sep='T')

    # FIXME: missing 'T' separator in last-modified
    md = Metadata(
        title=title,
        last_modified=today,
        version=ars_version,
        oscal_version=oscal_version
    )

    catalog = Catalog(
            uuid=uuid[9:],
            metadata=md,
            groups=groups,
        )
    root = Model(catalog=catalog).dict(by_alias=True, exclude_unset=True, exclude_none=True)
    print(json.dumps(root, indent=2, default=str))


if __name__ == "__main__":
    main()
